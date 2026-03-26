from __future__ import annotations

import csv
from datetime import date, datetime
import io
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook


@dataclass(slots=True)
class ParsedImportFile:
    source_type: str
    rows: list[dict]


REQUIRED_HEADERS = ["tradeDate", "symbol", "side", "price", "lots", "validity"]
REQUIRED_HEADER_LABELS = ["挂单时间", "股票代码", "方向", "委托价", "手数", "挂单方式"]
HEADER_ALIASES = {
    "挂单时间": "tradeDate",
    "股票代码": "symbol",
    "方向": "side",
    "委托价": "price",
    "手数": "lots",
    "挂单方式": "validity",
}


def _norm_header(value: str) -> str:
    return value.strip().replace(" ", "").lower()


def _map_header(value: str) -> str:
    return HEADER_ALIASES.get(_norm_header(value), _norm_header(value))


def _positive_float(value: str) -> float | None:
    try:
        parsed = float(value.replace(",", ""))
        return parsed if parsed > 0 else None
    except Exception:
        return None


def _positive_int(value: str) -> int | None:
    try:
        parsed = int(float(value.replace(",", "")))
        return parsed if parsed > 0 else None
    except Exception:
        return None


def _normalize_trade_date(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if text == "":
        return None
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return None


def _normalize_side(value: object) -> str | None:
    text = str(value or "").strip().upper()
    return text if text in {"BUY", "SELL"} else None


def _normalize_validity(value: object) -> str | None:
    text = str(value or "").strip().upper()
    return text if text in {"DAY", "GTC"} else None


def _build_row(record: dict, idx: int) -> dict:
    symbol = str(record.get("symbol", "")).strip().upper()
    side = _normalize_side(record.get("side", ""))
    price = _positive_float(str(record.get("price", "")).strip())
    lots = _positive_int(str(record.get("lots", "")).strip())
    validity = _normalize_validity(record.get("validity", ""))
    raw_trade_date = record.get("tradeDate")
    normalized_trade_date = _normalize_trade_date(raw_trade_date)
    trade_date_text = "" if raw_trade_date is None else str(raw_trade_date).strip()
    trade_date = normalized_trade_date or ""
    issues: list[str] = []
    if trade_date_text == "":
        issues.append("挂单时间不能为空，格式如 2026-03-25")
    elif not trade_date:
        issues.append("挂单时间格式需为 YYYY-MM-DD")
    if not symbol:
        issues.append("股票代码不能为空")
    if not side:
        issues.append("方向仅支持 BUY/SELL")
    if price is None:
        issues.append("委托价必须为正数")
    if lots is None:
        issues.append("手数必须为正整数")
    if not validity:
        issues.append("挂单方式仅支持 DAY/GTC")

    return {
        "rowNumber": idx,
        "tradeDate": trade_date or "",
        "symbol": symbol,
        "side": side or "BUY",
        "price": price or 0,
        "lots": lots or 0,
        "validity": validity or "DAY",
        "validationStatus": "ERROR" if issues else "VALID",
        "validationMessage": "；".join(issues) if issues else "校验通过",
    }


def _find_header_row(rows: list[list[str]]) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows):
        mapped_headers = [_map_header(cell) for cell in row]
        if all(required in mapped_headers for required in REQUIRED_HEADERS):
            return idx, mapped_headers
    raise ValueError(f"导入模板缺少必填列: {', '.join(REQUIRED_HEADER_LABELS)}")


def parse_import_file(
    file_name: str,
    content: bytes,
) -> ParsedImportFile:
    lower = file_name.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        raw_rows = [[str(cell or "") for cell in row] for row in csv.reader(io.StringIO(text))]
        header_idx, headers = _find_header_row(raw_rows)
        rows: list[dict] = []
        for raw in raw_rows[header_idx + 1 :]:
            if not raw or all(str(v).strip() == "" for v in raw):
                continue
            mapped = {}
            for idx, header in enumerate(headers):
                mapped[header] = "" if idx >= len(raw) or raw[idx] is None else str(raw[idx])
            rows.append(_build_row(mapped, len(rows) + 1))
        return ParsedImportFile(source_type="CSV", rows=rows)

    if lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            raise ValueError("Excel 文件中未找到工作表")
        raw_rows = [["" if cell is None else str(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        header_idx, headers = _find_header_row(raw_rows)

        rows: list[dict] = []
        for row in raw_rows[header_idx + 1 :]:
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            mapped = {}
            for idx, header in enumerate(headers):
                mapped[header] = "" if idx >= len(row) else row[idx]
            rows.append(_build_row(mapped, len(rows) + 1))
        return ParsedImportFile(source_type="XLSX", rows=rows)

    raise ValueError("当前仅支持 .csv 和 .xlsx 文件")


def build_import_template(target_trade_date: str | None = None) -> bytes:
    sample_trade_date = target_trade_date or "2026-03-25"
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(["A-share AI Trader 指令导入模板"])
    ws.append([f"请逐行填写挂单时间；方向支持 BUY/SELL，手数为正整数，挂单方式支持 DAY/GTC。"])
    ws.append(["挂单时间", "股票代码", "方向", "委托价", "手数", "挂单方式"])
    ws.append([sample_trade_date, "600519", "BUY", 1650, 1, "DAY"])
    ws.append([sample_trade_date, "000858", "SELL", 130, 3, "GTC"])

    guide = wb.create_sheet("填写说明")
    guide.append(["字段", "说明"])
    guide.append(["挂单时间", "每行单独填写，格式为 YYYY-MM-DD，例如 2026-03-25"])
    guide.append(["股票代码", "A 股 6 位股票代码，例如 600519"])
    guide.append(["方向", "仅支持 BUY / SELL"])
    guide.append(["委托价", "正数，最多保留两位小数"])
    guide.append(["手数", "正整数，1 手 = 100 股"])
    guide.append(["挂单方式", "支持 DAY / GTC；DAY 为当日挂单，GTC 为持续挂单"])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
